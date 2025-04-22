import pandas as pd
import geopandas as gpd
import numpy as np
from shapely.geometry import Point, LineString
from shapely import wkt
import polyline
# import pygsheets
import gspread
# import xlwings
import json
from gspread.exceptions import APIError

import time
from itertools import permutations
import copy

import os
from datetime import date
import openpyxl
import xlsxwriter
from config import *
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment


# # output in geojson

# # so all geometries in the right format

# #### useful general functions ####

def gspread_access_file_read_only(key, tab_list):
    """
    key = Google Sheets unique key in the URL
    title = name of the sheet you want to read
    returns a df of the sheet
    """
    gspread_creds = gspread.oauth(
        scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"],
        credentials_filename=client_secret_full_path,
        # authorized_user_filename=json_token_name,
    )
    list_of_dfs = []
       # # add in an exponential backoff 

    for tab in tab_list:
    
        # print(tab)
        wait_time = 5
        time.sleep(wait_time)
        gsheets = gspread_creds.open_by_key(key)
        # Access a specific tab
        spreadsheet = gsheets.worksheet(tab)
        
        # Attempt to fetch data from the sheet
        gsheets = gspread_creds.open_by_key(key)
        # Access a specific tab
        spreadsheet = gsheets.worksheet(tab)
        df = pd.DataFrame(spreadsheet.get_all_records(expected_headers=[]))
        df = df.replace('*', pd.NA).replace('Unknown', pd.NA).replace('--', pd.NA) # TODO maybe deal in another way?
        df = df.fillna('')
        list_of_dfs.append(df)

    if len(list_of_dfs) > 1: 
        # df = pd.concat(list_of_dfs, sort=False).reset_index(drop=True).fillna('')
        
        df = pd.concat(list_of_dfs, sort=False).reset_index(drop=True).fillna('')

    return df
 


# # #### useful geo functions ####

def geom_to_polyline(geom):
    if geom.geom_type == 'MultiLineString':
        # Convert each LineString within the MultiLineString to a polyline
        return [polyline.encode(list(line.coords)) for line in geom.geoms] # problem TypeError: 'MultiLineString' object is not iterable
    elif geom.geom_type == 'LineString':
        return polyline.encode(list(geom.coords))
    elif geom.geom_type == 'Point':
        return Point.encode(list(geom.coords)) # maybe? 
    else:
        raise ValueError("Unsupported geometry type")


def clean_dfs(df):
    # remove uknown, *, etc
    print(f'length of df at start of cleaned_dfs: {len(df)}')

    non_numeric_rows = []

    cleaned_df = df.copy()
    cleaned_df['Latitude'] = cleaned_df['Latitude'].astype(str)
    cleaned_df['Longitude'] = cleaned_df['Longitude'].astype(str)
    non_numeric_rows = cleaned_df[
        cleaned_df['Latitude'].str.contains(',') |
        cleaned_df['Longitude'].str.contains(',')].index.tolist()
    non_numeric_rows2 = cleaned_df[
        cleaned_df['Latitude'].str.contains('-') |
        cleaned_df['Longitude'].str.contains('-')].index.tolist()
            
    non_numeric_rows = non_numeric_rows + non_numeric_rows2 # + cleaned_df_missing
    # print(f'length of :{len(non_numeric_rows)}')
    cleaned_df = cleaned_df.drop(non_numeric_rows)
    # print("Rows to drop:", non_numeric_rows)
    # print("\nCleaned DataFrame:")
    # print(cleaned_df)
    
    # Create new DataFrames from the lists
    non_numeric_df = pd.DataFrame(non_numeric_rows)

    # print("\nNon-Numeric DataFrame:")
    # print(non_numeric_df)
    non_numeric_df.to_csv(f'{path_for_test_results}non_numeric_df_coords_{today_date}.csv')

                
    # print(f'length of df at end of cleand_dfs removed lat and long empty which is good because all point data:{len(cleaned_df)}')
    return cleaned_df

# def lat_lng_to_Point(df):

#     df = clean_dfs(df)
#     df = df.fillna('')


#     print('Successfully created a geometry column')


#     return df

def df_to_gdf(df, geometry_col, crs='EPSG:4326'):
    # Ensure the geometry column contains valid geometries
    # gdf = gpd.GeoDataFrame(df, geometry=gpd.GeoSeries.from_wkt(df[geometry_col]))
    # print(crs)
    gdf = gpd.GeoDataFrame(df, geometry=geometry_col, crs=crs)

    return gdf

def gdf_to_geojson(gdf, output_file):
    gdf.to_file(output_file, driver="GeoJSON")
    
    print(f"GeoJSON file saved to {output_file}")


def geojson_to_gdf(geojson_file):
    """
    convert to gdf
    """
    
    # Load GeoJSON file into a GeoDataFrame
    gdf = gpd.read_file(geojson_file)

    # get crs for final export
    crs = gdf.crs

    pipes_without_geo = gdf[gdf['geometry']==None]

    return gdf, crs


#### useful data check test functions ####

#### export functions ####

####
def get_standard_country_names():
    
    df = gspread_access_file_read_only(
        '1mtlwSJfWy1gbIwXVgpP3d6CcUEWo2OM0IvPD6yztGXI', 
        ['Countries'],
    )
    gem_standard_country_names = df['GEM Standard Country Name'].tolist()
    
    return gem_standard_country_names

gem_standard_country_names = get_standard_country_names()

def create_prep_file(key, tab):
    df = gspread_access_file_read_only(key, tab)

    df = df[df['acro'] != '']
    # Convert 'gspread_tabs' and 'sample_cols' to lists
    df['gspread_tabs'] = df['gspread_tabs'].apply(lambda x: x.split(';'))
    # df['sample_cols'] = df['sample_cols'].apply(lambda x: x.split(';'))
    df['gspread_tabs'] = df['gspread_tabs'].apply(lambda lst: [s.strip() for s in lst])
    # df['sample_cols'] = df['sample_cols'].apply(lambda lst: [s.strip() for s in lst])
    df.set_index('acro', inplace=True)
    return df


prep_df = create_prep_file(prep_file_key, prep_file_tab)


def create_conversion_df(key, tab):
    df = gspread_access_file_read_only(key, tab)
    # print(f'this is conversion df: {df}')
    
    df = df[['tracker', 'type', 'original units', 'conversion factor (capacity/production to common energy equivalents, TJ/y)']]
    df = df.rename(columns={'conversion factor (capacity/production to common energy equivalents, TJ/y)': 'conversion_factor', 'original units': 'original_units'})
    df['tracker'] = df['tracker'].apply(lambda x: x.strip())

    return df  

conversion_df = create_conversion_df(conversion_key, conversion_tab)  

def is_number(n):
    is_number = True
    try:
        num = float(n)
        # check for "nan" floats
        is_number = num == num   # or use `math.isnan(num)`
    except ValueError:
        is_number = False
    return is_number

def check_and_convert_int(x):
    if is_number(x):
        return int(x)
    else:
        return np.nan

def check_and_convert_float(x):
    if is_number(x):
        return float(x)
    else:
        return np.nan
    
def find_missing_coords(df):

    df['float_col_clean_lat'] = df['Latitude'].apply(lambda x: check_and_convert_float(x))
    
    df['float_col_clean_lng'] = df['Longitude'].apply(lambda x: check_and_convert_float(x))

    # print("Check which rows will be dropped because nan coords:")
    # print(df[df.isna().any(axis=1)])
    nan_df = df[df.isna().any(axis=1)]
    # nan_df.to_csv(f'{path_for_test_results}{df["tracker"].loc[0]}_nan_coords_{today_date}.csv')

    df = df.dropna(subset = ['float_col_clean_lat', 'float_col_clean_lng'], how = 'any')

    return df
    
def find_missing_cap(df):
    
    # GOGET does not have capacity 
    if df['tracker'].loc[0] == 'GOGET':
        pass
    else:
        df['float_col_clean_cap'] = df['capacity'].apply(lambda x: check_and_convert_float(x))
        
        # print("Check which rows will be dropped because nan capacity:")
        # print(df[df.isna().any(axis=1)])
        nan_df = df[df.isna().any(axis=1)]
        # nan_df.to_csv(f'{path_for_test_results}{df["tracker"].loc[0]}_nan_capacity_{today_date}.csv')
        df = df.dropna(subset = ['float_col_clean_cap'])
        
    return df

def convert_coords_to_point(df):
    crs = 'EPSG: 4326'
    geometry_col = 'geometry'
    for row in df.index:
        df.loc[row,'geometry'] = Point(df.loc[row,'Longitude'], df.loc[row,'Latitude'])
    gdf = gpd.GeoDataFrame(df, geometry=geometry_col, crs=crs)
    
    return gdf

def update_previous_about_page(prev_key, new_release_date):
    gspread_creds = gspread.oauth(
            scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"],
            credentials_filename=client_secret_full_path,
            # authorized_user_filename=json_token_name,
        )
    wait_time = 5
    time.sleep(wait_time)
    gsheets = gspread_creds.open_by_key(prev_key)
        # List all sheet names
    sheet_names = [sheet.title for sheet in gsheets.worksheets()]
    print(f"Sheet names previous release:", sheet_names)
    # adjust About Africa Energy Tracker tab with new release date
    aet_about = sheet_names[0]
    aet_about = gsheets.worksheet(aet_about) 

    aet_data = pd.DataFrame(aet_about.get_all_records(expected_headers=[]))
    updated_aet_data = aet_data.replace('July 2024', 'August 2024')
    
    updated_aet_data.to_excel(f'{path_for_test_results}aet_about_page{today_date}.xlsx')
    print(f'Pause to check test results folder for updated AET about page data.')
    input("Press Enter to continue...")
    
    return updated_aet_data

def find_about_page(tracker,key):
<<<<<<< HEAD

=======
>>>>>>> 81808373619ec7caaa2f0f4b2814984378a66e84
        gspread_creds = gspread.oauth(
            scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"],
            credentials_filename=client_secret_full_path,
            # authorized_user_filename=json_token_name,
        )
        wait_time = 5
        time.sleep(wait_time)
        gsheets = gspread_creds.open_by_key(key)
            
        # List all sheet names
        sheet_names = [sheet.title for sheet in gsheets.worksheets()]
        print(f"{tracker} Sheet names:", sheet_names)
        # Access a specific sheet by name
        first_tab = sheet_names[0]
        first_sheet = gsheets.worksheet(first_tab)  # Access the first sheet
        
        last_tab = sheet_names[-1]
        last_sheet = gsheets.worksheet(last_tab)  # Access the last sheet

        # print("First sheet name:", sheet.title)
        if 'About' not in first_sheet.title:
            print('Looking for about page in last tab now, first one no.')
            # handle for goget and ggit, goit who put it in the last tab
            if 'About' not in last_sheet.title:
                if 'Copyright' not in last_sheet.title:
                    print('Checked first and last tab, no about page found not even for copyright. Pausing.')
                    input("Press Enter to continue...")
                else:
                    print(f'Found about page in last tab: {last_tab}')
                    sheet = last_sheet
            else:
                print(f'Found about page in last tab: {last_tab}')
                sheet = last_sheet
        else:
            print(f'Found about page in first tab: {first_tab}')
            sheet = first_sheet
        
        # TODO why does GCPT and GCTT not pull in all data? formatting? GCMT was fine and merged.
        data = pd.DataFrame(sheet.get_all_records(expected_headers=[]))
        about_df = data.copy()
    
        return about_df
    
def write_to_about_page_file(about_page_df_dict):

        # I should update this so that we can have the full data download
        about_output = f'{path_for_data_dwnld}{today_date}_about_pages.xlsx'        
        # pull out previous overall about page - done in dict
        
        # Write each DataFrame to a separate sheet

        if os.path.exists(about_output):
            
            # try:
            with pd.ExcelWriter(about_output, engine='xlsxwriter', mode='a') as writer:
                for sheet_name, df in about_page_df_dict.items():
                    print(sheet_name)
                    print(len(df))
                    try:
                        df.to_excel(writer, sheet_name=sheet_name, index=False) # if_sheet_exists='replace'
                    except AttributeError as e:
                        print(f'problem on {sheet_name}: {e}')
        else:
            with pd.ExcelWriter(about_output, engine='xlsxwriter') as writer:
                for sheet_name, df in about_page_df_dict.items():
                    print(sheet_name)
                    try:
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                    except AttributeError as e:
                        print(f'problem on {sheet_name}: {e}')
        # except FileNotFoundError:
        #     with pd.ExcelWriter(about_output, engine='openpyxl') as writer:
        #         for sheet_name, df in about_page_df_dict.items():
        #             # print(sheet_name)
        #             try:
        #                 df.to_excel(writer, sheet_name=sheet_name, index=False)
        #             except AttributeError as e:
        #                 print(f'problem on {sheet_name}: {e}')
        # finally:
        #     print(f'this always runs.')
<<<<<<< HEAD
=======

>>>>>>> 81808373619ec7caaa2f0f4b2814984378a66e84
            
        return about_output

def find_region_colname(df):
    continent_list = ['Africa', 'Americas', 'Oceania', 'Asia', 'Europe']
    cols = df.columns
    col_reg_name = '' # initialize 
    for col in cols:
        # random row
        random = int(len(df) / 2)
        # print(random)
        
        if df[col].loc[random] in continent_list:
            col_reg_name = col 
            print(f'this is region col: {col_reg_name} to filter on!')
             
    return col_reg_name    


# def col_differences(df):
    # todo pull over what you have in the other script for release notes to help highlight code changes
    # needed based on column name changes 
   
<<<<<<< HEAD
=======

>>>>>>> 81808373619ec7caaa2f0f4b2814984378a66e84

def create_all_dfs(df):
    all_dict = df.to_dict(orient='index')
    # print(all_dict)
    list_of_gdfs = [] # map 
    list_of_dfs = [] # data download 
    about_page_df_dict = {} # tracker: sheetname: tracker_df_tabs (pull out )
    
    for key, value in all_dict.items():
        print(f'we are on tracker: {key}')
        # we could read in the sheet that has the release date that I manage 15l2fcUBADkNVHw-Gld_kk7EaMiFFi8ysWt6aXVW26n8 

        if local_copy == False:

            # print(all_dict[key]['gspread_tabs'])
            # dfs_about_dict = pd.read_excel(all_dict[key]['gspread_key'], sheet_name=None)
            about_page_df = find_about_page(key, all_dict[key]['gspread_key'])
            about_page_df_dict[f'About {key}'] = about_page_df
            
            # get the df for each acro key
            df = gspread_access_file_read_only(all_dict[key]['gspread_key'], all_dict[key]['gspread_tabs'])
            # find all col name changes 
            col_info = {}

            for col in df.columns:                
                if 'ref' in col.lower():
                    print('ref found pass')
                else:  
                    col_types = [set(df[col].to_list())] # len can be a lot, cap at 30
                    col_values = [df[col].dtype] # len = 1
                    col_range = [df[col].min(), df[col].max()] # len = 2
                    col_mid = df[col].mean() # len = 1
                    col_sum = df[col].sum()   # len = 1
                    # cap for readability 
                    for colinfo in [col_types, col_values, col_range, col_mid, col_sum]:
                        if len(colinfo) >30:
                            colinfo = colinfo[:30]
                    col_info[col] = {'col_values': set(df[col].to_list()), 'col_type': df[col].dtype}
                
            col_info_df = pd.DataFrame(col_info)
            # col_info_df = col_info_df.T # transpose
            col_info_df.to_csv(f'test_results/{key}_column_info_{iso_today_date}.csv')
            print(f'saved col info for {key} here: "test_results/{key}_column_info_{iso_today_date}.csv"')
            
            df['tracker'] = key
            
            # filter by Africa, but GOGET names it GEM region .. key error?
            col_reg_name = find_region_colname(df)
            df = df[df[col_reg_name] == 'Africa'] 
            
            df = df.fillna('')
            
            col_info_df = pd.DataFrame(col_info)
            # col_info_df = col_info_df.T # transpose
            col_info_df.to_csv(f'test_results/{key}_column_info_{iso_today_date}.csv')
            print(f'saved col info again for {key} here: "test_results/{key}_column_info_{iso_today_date}.csv"')
             
            
            # clean up missing coords, # do latercapacity, status, country
            df = find_missing_coords(df)
            # handle lat lng into geometry so when concat all gdfs already
            # convert df to gdf

            gdf = convert_coords_to_point(df)
            # append df to list of dfs for data download
            list_of_dfs.append(df)
            df.to_excel(f'{path_for_test_results}{key}_df_{today_date}.xlsx', index=False)

            # append gdf to list of gdfs for map - though now we can have it as a csv for faster AET non tile load
            list_of_gdfs.append(gdf)
            print(f'added {key}')
            
            # here let's save the gdfs you pulled from the gspreadsheet since that's time consuming, local copy
            gdf_to_geojson(gdf, f'{path_for_test_results}{key}_gdf_{today_date}.geojson')
        
        
        else:
            # local copy
            for file in os.listdir(path_for_test_results):
                if file.endswith(".geojson") & (key in file):
                    gdf = gpd.read_file(f'{path_for_test_results}{file}')
                    list_of_gdfs.append(gdf)
                    print(f'added {file} to list of gdfs')
                elif file.endswith(".xlsx") & (key in file):
                    df = pd.read_excel(f'{path_for_test_results}{file}')
                    list_of_dfs.append(df)
                    print(f'added {file} to list of dfs')                    
    
    
    prev_about_page_df = update_previous_about_page(prev_key, new_release_date)
    about_page_df_dict['About Africa Energy Tracker'] = prev_about_page_df             
    # about_page_dfs = pd.DataFrame(about_page_df_dict, index=False)
    # about_page_dfs.to_excel(f'{path_for_data_dwnld}{today_date}about_pages.xlsx')
    
    # about_output = write_to_about_page_file(about_page_df_dict)
    about_output = 'to do'
    
    return list_of_gdfs, list_of_dfs, about_output

list_of_gdfs, list_of_dfs, about_output = create_all_dfs(prep_df)

# add into list of dfs, GGIT and GOIT 

# def get_about_page_prev(prev_key):
#     prev_key = prev_key
    

def incorporate_geojson_trackers(GOIT, GGIT, GGIT_lng, list_of_gdfs):
    
    pipes_gdf = gpd.read_file(GOIT)
    pipes_gdf['tracker'] = 'GOIT'
    ggit_gdf = gpd.read_file(GGIT)
    ggit_gdf['tracker'] = 'GGIT'
    ggit_lng_gdf = gpd.read_file(GGIT_lng)
    ggit_lng_gdf['tracker'] = 'GGIT-lng'

    list_of_gdfs.append(pipes_gdf)
    list_of_gdfs.append(ggit_gdf)
    list_of_gdfs.append(ggit_lng_gdf)

    # to do, we need to convert gdf to df so the data download has all the data
<<<<<<< HEAD
    return list_of_gdfs
=======

>>>>>>> 81808373619ec7caaa2f0f4b2814984378a66e84


list_of_gdfs = incorporate_geojson_trackers(goit_geojson, ggit_geojson, ggit_lng_geojson, list_of_gdfs)

                

def create_data_download(list_of_gdfs, list_of_dfs, about_output):
    list_of_gdfs = list_of_gdfs
    list_of_dfs = list_of_dfs

    # create file name xlsx writer
    xlsfile = f'{path_for_data_dwnld}{today_date}africa-energy-tracker-data.xlsx'
    # # wb = load_workbook(xlsfile)
    # wb = xlsxwriter.Workbook(xlsfile)

    # # Add a new sheet
    # # wb = wb.add_worksheet(title=f"{tracker}")

    # new_sheet = wb.create_sheet(title=f"{tracker}")
    # wb.save(xlsfile)
    # wb.close()
    with pd.ExcelWriter(xlsfile, engine='openpyxl') as writer:
        # for df in list choose a tab name from tracker 
        # print entire df into tab
        # for pipelines us gdf
        for df in list_of_dfs:
            try:
                tracker = df['tracker'].loc[0]
                print(len(df))
            except KeyError as e:
                print(f"Error: Not a tracker sheet. {e}")
            else:
                print(f"Result works for {tracker}")
                print(f'Adding this tracker to dt dwnld: {tracker}')
                # filter by Africa, but GOGET names it GEM region .. key error?
                # do this above when making list of dfs and gdfs
                # col_reg_name = find_region_colname(df)
                # df_africa = df[df[col_reg_name] == 'Africa'] 
                # df_africa.to_excel(writer, sheet_name=f'{tracker}', index=False)
                columns_to_drop = ['tracker', 'float_col_clean_lat', 'float_col_clean_lng']

                # Check which columns to drop actually exist in the DataFrame
                existing_columns_to_drop = [col for col in columns_to_drop if col in df.columns]

                # Drop the existing columns
                if existing_columns_to_drop:
                    df = df.drop(columns=existing_columns_to_drop)
                df.to_excel(writer, sheet_name=f'{tracker}', index=False)
                print(f"DataFrame {tracker} written to {xlsfile}")
            finally:
                print("This will always run.")
                
    print(f"DataFrames written to {xlsfile}")
    
    # TODO 
    # merge this file with the one with about page stuff
    about_output # file name to merge with
  
# create_data_download(list_of_gdfs, list_of_dfs, about_output) 
<<<<<<< HEAD
=======

>>>>>>> 81808373619ec7caaa2f0f4b2814984378a66e84


def split_goget_ggit(list_of_gdfs):
    custom_list_of_gdfs = []
    # print(f'this is length before: {len(list_of_gdfs)}')
    # Read the Excel file

    
    for gdf in list_of_gdfs:
        # print(df['tracker'])
        # df = df.fillna('') # df = df.copy().fillna("")
        if gdf['tracker'].iloc[0] == 'GOGET':
            # print(gdf.columns)
            # oil
            gdf = gdf.copy()
            # df_goget_missing_units.to_csv('compilation_output/missing_gas_oil_unit_goget.csv')
            # gdf['tracker_custom'] = gdf.apply(lambda row: 'GOGET - gas' if row['Production - Gas (Million m³/y)'] != '' else 'GOGET - oil', axis=1)
            gdf['tracker_custom'] = 'GOGET - oil'
            custom_list_of_gdfs.append(gdf)
            
        elif gdf['tracker'].iloc[0] == 'GGIT-lng':
            gdf_ggit_missing_units = gdf[gdf['FacilityType']=='']
            # df_ggit_missing_units.to_csv('compilation_output/missing_ggit_facility.csv')
            gdf = gdf[gdf['FacilityType']!='']
            gdf['tracker_custom'] = gdf.apply(lambda row: 'GGIT - import' if row['FacilityType'] == 'Import' else 'GGIT - export', axis=1)
            # print(gdf[['tracker_custom', 'tracker', 'FacilityType']])
        
            custom_list_of_gdfs.append(gdf)

        else:
            gdf['tracker_custom'] = gdf['tracker']
        
            custom_list_of_gdfs.append(gdf)
                
    # print(f'this is length after: {len(custom_list_of_gdfs)}')
    # for df in custom_list_of_dfs:
    #     if df['tracker'].iloc[0] == 'GOGET':
    #         print(df['tracker_custom'])
    return custom_list_of_gdfs

custom_list_of_gdfs = split_goget_ggit(list_of_gdfs)


def assign_conversion_factors(list_of_gdfs, conversion_df):
    # add column for units 
    # add tracker_custom
    augmented_list_of_gdfs = []
    for gdf in list_of_gdfs:
        gdf = gdf.copy()
        # print(gdf['tracker'].loc[0])
        # print(gdf['geometry'])
        # print(type(df))
        # print(df.columns.to_list()) # it's saying tracker is equalt to custom_tracker for ggit-lng shouldn't be the case 

        if gdf['tracker'].iloc[0] == 'GOGET': 
            # print(f'We are on tracker: {gdf["tracker"].iloc[0]} length: {len(gdf)}')

            for row in gdf.index:
                if gdf.loc[row, 'tracker_custom'] == 'GOGET - oil':
                    gdf.loc[row, 'original_units'] = conversion_df[conversion_df['tracker']=='GOGET - oil']['original_units'].values[0]
                    gdf.loc[row, 'conversion_factor'] = conversion_df[conversion_df['tracker']=='GOGET - oil']['conversion_factor'].values[0]
                # elif gdf.loc[row, 'tracker_custom'] == 'GOGET - gas':  
                #     gdf.loc[row, 'original_units'] = conversion_df[conversion_df['tracker']=='GOGET - gas']['original_units'].values[0]
                #     gdf.loc[row, 'conversion_factor'] = conversion_df[conversion_df['tracker']=='GOGET - gas']['conversion_factor'].values[0]

            gdf['tracker'] = 'GOGET'
     
            augmented_list_of_gdfs.append(gdf)
        elif gdf['tracker'].iloc[0] == 'GGIT-lng':
            # print(f'We are on tracker: {gdf["tracker"].iloc[0]} length: {len(gdf)}')
            for row in gdf.index:
                if gdf.loc[row, 'tracker_custom'] == 'GGIT - export':
                    gdf.loc[row, 'original_units'] = conversion_df[conversion_df['tracker']=='GGIT - export']['original_units'].values[0]
                    gdf.loc[row, 'conversion_factor'] = conversion_df[conversion_df['tracker']=='GGIT - export']['conversion_factor'].values[0]
                elif gdf.loc[row, 'tracker_custom'] == 'GGIT - import':  
                    gdf.loc[row, 'original_units'] = conversion_df[conversion_df['tracker']=='GGIT - import']['original_units'].values[0]
                    gdf.loc[row, 'conversion_factor'] = conversion_df[conversion_df['tracker']=='GGIT - import']['conversion_factor'].values[0]
            # now let's rename GGIT tracker so lng is not needed
            # df['tracker'] = 'GGIT'
            # print(gdf['tracker'])
            # print(gdf['tracker_custom'])
            augmented_list_of_gdfs.append(gdf)

        else:
            # print(f'We are on tracker: {gdf["tracker"].iloc[0]} length: {len(gdf)}')
            # This should apply to all rows not just one.
            gdf['tracker_custom'] = gdf["tracker"].iloc[0]
            gdf['original_units'] = conversion_df[conversion_df['tracker']==gdf['tracker'].iloc[0]]['original_units'].values[0]
            gdf['conversion_factor'] = conversion_df[conversion_df['tracker']==gdf['tracker'].iloc[0]]['conversion_factor'].values[0]
      
            augmented_list_of_gdfs.append(gdf)

    return augmented_list_of_gdfs

list_of_gdfs_with_conversion = assign_conversion_factors(custom_list_of_gdfs, conversion_df)


def rename_dfs(list_of_dfs_with_conversion):
    # renaming_cols_dict['GGIT -lng'] = {renaming dict}
    # renaming_dict = {} # GOGPT: {col:newcol, col2:newcol2}
    renamed_dfs = []
    all_trackers = []
    len_africa = 0 # initialize
    # print(f'Length of original gdfs: {len(list_of_dfs_with_conversion)}')
    # print(list_of_gdfs)
    for df in list_of_dfs_with_conversion:
        # print(item.columns)
        # print(item['tracker'])
        # print(item['tracker_custom'])
        tracker_sel = df['tracker'].iloc[0] # GOGPT, GGIT, GGIT-lng, GOGET
        # print(tracker_sel)
        
        # tracker_sel = tracker_sel.split('-')[0]
        all_trackers.append(tracker_sel)
        renaming_dict_sel = renaming_cols_dict[tracker_sel]
        df = df.rename(columns=renaming_dict_sel) 
        # print(f"Len of df name col: {len(df['name'])}")
        # print(f"Len of df region africa:{len(df[df['region']=='Africa'])}")
        len_africa += (len(df[df['region']=='Africa']))

        # df = find_missing_cap(df) # no need them will fill in average ... still remember for GOGET to differentiate use cap-gas if custom_tracker gas etc

        renamed_dfs.append(df)
    # print(f'all tracker names: {all_trackers}')
    # print(f'Length of renamed gdfs: {len(renamed_dfs)}')
    
    # drop un needed cols with final_cols  

    # sort 
    # drop duplicate on unit id
    # concat all gdfs 
    one_df = pd.concat(renamed_dfs, sort=False).reset_index(drop=True)
    one_df = one_df.drop_duplicates(subset=['id', 'geometry'])
    
    
    
    # print(f'This is oned_df columns lets see if there are two names: {one_df.columns}')
    # df = one_df[final_cols]
    df = one_df.copy()
    # print(df.columns)
    df.to_csv(f'{path_for_test_results}concatted_df_{today_date}.csv')
<<<<<<< HEAD
=======

>>>>>>> 81808373619ec7caaa2f0f4b2814984378a66e84
    
    return df

concatted_gdf = rename_dfs(list_of_gdfs_with_conversion)


# def harmonize_countries(gdf):
#     gdf = gdf.copy()


#     """
#     Standardize country names, based on file "GEM Country Naming Conventions"
#     """

#     country_col = 'areas'
        
#     # strip white space
#     gdf[country_col] = gdf[country_col].str.strip()
        
#     # harmonize countries:
#     country_harm_dict = {
#         'Czechia': 'Czech Republic',
#         'Ivory Coast': "Côte d'Ivoire",
#         "Cote d'Ivoire": "Côte d'Ivoire", # adds accent
#         "Republic of Congo": "Republic of the Congo", # adds "the"
#         "Rep Congo": "Republic of the Congo",
#         "Democratic Republic of Congo": "DR Congo",
#         "Democratic Republic of the Congo": "DR Congo", # in case step above adds "the"
#         "Republic of Guinea": "Guinea",
#         "Republic of Sudan": "Sudan",
#         "FYROM": "North Macedonia",
#         "Chinese Taipei": "Taiwan",
#         "East Timor": "Timor-Leste",
#         "USA": "United States",
#         'Turkey': 'Türkiye',
#         'Canary Islands': 'Spain', 
#     }
#     for key in country_harm_dict.keys():
#         sel = gdf[gdf[country_col]==key]
#         if len(sel) > 0:
#             print(f"Found non-standardized country name before trying to standardize: {key} ({len(sel)} rows)")
#         gdf[country_col] = gdf[country_col].str.replace(key, country_harm_dict[key])
        
#     # clean up, checking if countries are in standard GEM list
#     hyphenated_countries = ['Timor-Leste', 'Guinea-Bissau']
#     for row in gdf.index:
#         if pd.isna(gdf.at[row, country_col])==False:    
#             try:
#                 countries_list = gdf.at[row, country_col].split(', ')
#                 countries_list = [x.split('-') for x in countries_list if x not in hyphenated_countries]
#             except:
#                 print("Error!" + f" Exception for row {row}, country_col: {gdf.at[row, country_col]}")
#                 countries_list = []
                
#             # flatten list
#             countries_list = [
#                 country
#                 for group in countries_list
#                 for country in group
#             ]
#             # clean up
#             countries_list = [x.strip() for x in countries_list]
        
#             # check that countries are standardized
#             for country in countries_list:
#                 if country not in gem_standard_country_names:
#                     print(f"For row {row}, non-standard country name after trying to standardize: {country}")
#         else:
#             continue
#             # print(f"No countries listed for row {row}")


#     return gdf

# harmonized_gdf = harmonize_countries(concatted_gdf)

# def handle_multiple_countries(gdf):
#     # replace separate , with ;
#     # parse out so that area1 has an appropriate africa_countries
#     gdf = gdf.copy()
    
#     country_exceptions = ['Saint Helena, Ascension and Tristan da Cunha (UK)', 'Guinea-Bissau']
    
#     # print(f"List of areas before handle mult countries: {set(gdf['areas'].to_list())}")
#     # # goget
#     # gdf['areas'] = gdf['areas'].fillna('')
#     # # assign first one to area like all other countries 
#     # gdf['area'] = gdf.apply(lambda row: row['areas'] if row['areas'] in country_exceptions or '-' not in row['areas'] else row['areas'].split('-')[0], axis=1)
#     # # assign second one to new column so can merge together later if needed for goget list
#     # gdf['area_extra'] = gdf.apply(lambda row: row['areas'].split('-')[1] if '-' in row['areas'] and row['areas'] not in country_exceptions else '', axis=1)
    
#     # # pipelines 
#     # # assign first one to area like all other countries 
#     # # assign actual coutnry if its in exceptions otherwise split on -
#     # gdf['area'] = gdf.apply(lambda row: row['areas'] if row['areas'] in country_exceptions else row['areas'].split('-')[0], axis=1)
#     # # assign actual coutnry declared in previous line if its in exceptions, otherwise split on ,

#     # gdf['area'] = gdf.apply(lambda row: row['area'] if row['area'] in country_exceptions else row['areas'].split(',')[0], axis=1)

#     # now make it so that areas will be a good list of them all separated by ;
#     gdf['areas'] = gdf['areas'].apply(lambda x: x if x in country_exceptions else x.strip().replace('-',';'))
#     gdf['areas'] = gdf['areas'].apply(lambda x: x if x in country_exceptions else x.strip().replace(',',';'))
#     gdf['areas'] = gdf['areas'].apply(lambda x: x.replace('; ', ';'))
#     # print(f"List of areas after handle mult countries: {set(gdf['areas'].to_list())}")
#     return gdf

# concatted_gdf = handle_multiple_countries(harmonized_gdf)

def filter_by_country_all_gdfs(gdf):
    # df = pd.read_csv(file_path) # file_path in config file
    print(f'length of df at start of filter by regino: {len(gdf)}')
    
    # df['areas'] = df['areas'].apply(lambda x: x.replace('Guinea,Bissau','Guinea-Bissau'))

    # df['areas'] = df['areas'].apply(lambda x: x.split(','))
    # df['areas'] = df[df['tracker']=='GOGET']['areas'].apply(lambda x: x.split('-'))
    
    # africa_df1 = gdf[gdf['area'].isin(africa_countries)] 
    # don't need this anymore because should have all fixed so no need to use region 
    # africa_df2 = gdf[gdf['area1'].isin(africa_countries)]
    
    # need this to remove pipeline regions that are not in africa
<<<<<<< HEAD
=======

>>>>>>> 81808373619ec7caaa2f0f4b2814984378a66e84
    africa_df = gdf[gdf['region']=='Africa']
    # africa_df = pd.concat([africa_df1, africa_df2])
    # africa_df = pd.concat([africa_df, africa_df3])
    # print(f'filter by country before drop dup on id: {len(africa_df1)}')

    # africa_df = africa_df1.drop_duplicates(subset=['id', 'name']) # TODO check
    # africa_df = df['areas'].apply_row(lambda x: row.drop() if x not in africa_countries else row) 
    # print(f'Africa filter by country: {len(africa_df)}')
    print(f'length of df after of filter by region: {len(gdf)}')

    return africa_df

africa_gdf = filter_by_country_all_gdfs(concatted_gdf)



def remove_null_geo(concatted_gdf):

    # print(set(concatted_gdf['geometry'].to_list()))
    # print(f'length of df at start of remove_null_geo: {len(concatted_gdf)}')
    # concatted_gdf = concatted_gdf[concatted_gdf['geometry']!='null']
    good_keywords = ['point', 'line']
    filtered_geo = concatted_gdf[concatted_gdf['geometry'].apply(lambda x: any(keyword in str(x).lower() for keyword in good_keywords))]
    # print(f'length of df at after filter for point and line geo: {len(filtered_geo)}')
    dropped_geo = pd.concat([concatted_gdf, filtered_geo], ignore_index=True).drop_duplicates(keep=False)
    # print(dropped_geo)
    # print(dropped_geo[['tracker', 'name', 'geometry']])
    
    return filtered_geo

africa_gdf_cleaned = remove_null_geo(africa_gdf)


# def analyze_geojson(filepath):
#     gdf = gpd.read_file(filepath)
#     # print(len(gdf))
#     # print(gdf)
#     # print(gdf.columns)
#     df = pd.DataFrame(gdf)
#     # print(len(df))

#     good_keywords = ['point', 'line']
#     filtered_geo = df[df['geometry'].apply(lambda x: any(keyword in str(x).lower() for keyword in good_keywords))]
#     # print(len(filtered_geo))
#     dropped_geo = pd.concat([df, filtered_geo], ignore_index=True).drop_duplicates(keep=False)
#     # print(dropped_geo)
#     # print(dropped_geo[['name', 'geometry']])
#     return gdf
    
def conversion_multiply(row):
    cap = float(row['cleaned_cap'])
    factor = float(row['conversion_factor'])
    # print(f'this is factor! {factor}')
    result = float(cap * factor)
    # print(f'this is result! {result}')
    return result

def workaround_no_sum_cap_project(gdf):
    gdf = gdf.copy()
    
    # result = int()
    
    # group by id
    # summed cleaned cap
    # that's the project cap
    project_cap_df = gdf.groupby('name', as_index=False)['capacity'].sum()
    print(f'this is cols of project_cap_df: {project_cap_df.columns}')
    project_cap_df = project_cap_df.rename(columns={'capacity': 'capacity-details'})
    
    # merge on name to gdf
    
    gdf = pd.merge(left=gdf, right=project_cap_df, on='name', how='outer')
    
    return gdf

def workaround_display_cap(row):
    cap = row['capacity'] 
    units_of_m = str(row['original_units'])
    if isinstance(cap, (int, float)):
        cap = str((round(cap, 3))) # handle rounding and converting from string to float to round later 
        result = f'{cap} {units_of_m}'
    else:
        result = ''
    return result

def workaround_display_cap_total(row):
    cap = row['capacity-details']
    units_of_m = str(row['original_units'])
    if isinstance(cap, (int, float)):
        cap = str(float(round(cap, 3)))
        result = f'{cap} ({units_of_m})'
    else:
        result = ''
    return result


def capacity_conversions(gdf): 

# you could multiply all the capacity/production values in each tracker by the values in column C, 
# "conversion factor (capacity/production to common energy equivalents, TJ/y)."
# For this to work, we have to be sure we're using values from each tracker that are standardized
# to the same units that are stated in this sheet (column B, "original units").

# need to get nuances for each tracker, it's just GGIT if point then do lng terminal if line then do pipeline

    gdf_converted = gdf.copy()
    
    # first let's get GHPT cap added 
    # print(len(gdf_converted))
    ghpt_only = gdf_converted[gdf_converted['tracker']=='GHPT'] # for GGPT we need to re run it to get it 
    gdf_converted = gdf_converted[gdf_converted['tracker']!='GHPT']
    ghpt_only['capacity'] = ghpt_only.apply(lambda row: row['capacity1'] + row['capacity2'], axis=1)
    gdf_converted = pd.concat([gdf_converted, ghpt_only],sort=False).reset_index(drop=True)
    # print(len(gdf_converted))
    
    # # second let's handle GOGET -- need to differentiate use cap-gas if custom_tracker gas etc
    # goget_oil_only = gdf_converted[gdf_converted['tracker_custom']=='GOGET - oil'] # for GGPT we need to re run it to get it 
    # goget_gas_only = gdf_converted[gdf_converted['tracker_custom']=='GOGET - gas'] # for GGPT we need to re run it to get it 

    # goget_oil_only['capacity'] = goget_oil_only['capacity_oil']
    # goget_gas_only['capacity'] = goget_gas_only['capacity_gas']

    # gdf_converted = gdf_converted[gdf_converted['tracker']!='GOGET']
    # gdf_converted = pd.concat([gdf_converted, goget_oil_only],sort=False).reset_index(drop=True)
    # gdf_converted = pd.concat([gdf_converted, goget_gas_only],sort=False).reset_index(drop=True)

    

    gdf_converted['cleaned_cap'] = pd.to_numeric(gdf_converted['capacity'], errors='coerce')

    total_counts_trackers = []
    avg_trackers = []

    print(f"this is all trackers: {set(gdf_converted['tracker'].to_list())}")
    
    gdf_converted['tracker'] = gdf_converted['tracker'].fillna('')
    print(f"this is all trackers: {set(gdf_converted['tracker'].to_list())}")
    
    gdf_converted = gdf_converted[gdf_converted['tracker']!=''] # new to filter out nan


    for tracker in set(gdf_converted['tracker'].to_list()):
        total = len(gdf_converted[gdf_converted['tracker'] == tracker])
        sum = gdf_converted[gdf_converted['tracker'] == tracker]['cleaned_cap'].sum()
        avg = sum / total
        total_pair = (tracker, total)
        total_counts_trackers.append(total_pair)
        avg_pair = (tracker, avg)
        avg_trackers.append(avg_pair)

    # print(total_counts_trackers)
    # print(avg_trackers)
 
    # print(total_counts_trackers)
    # print(avg_trackers) 
    
    for row in gdf_converted.index:
        cap_cleaned = gdf_converted.loc[row, 'cleaned_cap']
        tracker = gdf_converted.loc[row, 'tracker']
        if pd.isna(cap_cleaned):
            for pair in avg_trackers:
                if pair[0] == tracker:
                    gdf_converted.loc[row, 'cleaned_cap'] = pair[1]
        cap_cleaned = gdf_converted.loc[row, 'cleaned_cap']
        if pd.isna(cap_cleaned):
            print('still na')
 

    pd.options.display.float_format = '{:.0f}'.format
    gdf_converted['scaling_capacity'] = gdf_converted.apply(lambda row: conversion_multiply(row), axis=1)
    # print(f"{gdf_converted['scaling_capacity'].min().round(2)}")
    # print(f"{gdf_converted['scaling_capacity'].max().round(2)}")
    # print(f"scaling_cap stats: {gdf_converted['scaling_capacity'].describe()}")
    gdf_converted['capacity-table'] = gdf_converted.apply(lambda row: workaround_display_cap(row), axis=1)

    gdf_converted = workaround_no_sum_cap_project(gdf_converted) # adds capacity-details 
    
    gdf_converted['capacity-details'] = gdf_converted.apply(lambda row: workaround_display_cap_total(row), axis=1)
    # gdf_converted['capacity-plant-total'] = gdf_converted.apply(lambda row: workaround_no_sum_cap_project(row), axis=1)

    return gdf_converted

africa_df_converted = capacity_conversions(africa_gdf_cleaned)

def fix_status_inferred(df):
    df = df.fillna('')
    # print(f"Statuses before: {set(df['status'].to_list())}")
    inferred_statuses_cancelled = df['status'].str.contains('cancelled - inferred')
    inferred_statuses_shelved = df['status'].str.contains('shelved - inferred')
    
    # # print(inferred_statuses_cancelled['status']!=False)
    # # print(len(inferred_statuses_shelved))
    df.loc[inferred_statuses_cancelled, 'status'] = 'cancelled'
    df.loc[inferred_statuses_shelved,'status'] = 'shelved'
    # for row in df.index:
    #     if 'shelved - inferred' in df.loc[row, 'status']:
    #         df.loc[row, 'status'] = 'shelved'
    
    # print(f"Statuses after: {set(df['status'].to_list())}")

    return df

def map_ready_statuses(gdf):
    gdf_map_ready = gdf.copy()
  
    gdf_map_ready = fix_status_inferred(gdf_map_ready)

    gdf_map_ready['status_legend'] = gdf_map_ready.copy()['status'].str.lower().replace({
                # proposed_plus
                'proposed': 'proposed_plus',
                'announced': 'proposed_plus',
                'discovered': 'proposed_plus',
                # pre-construction_plus
                'pre-construction': 'pre-construction_plus',
                'pre-permit': 'pre-construction_plus',
                'permitted': 'pre-construction_plus',
                # construction_plus
                'construction': 'construction_plus',
                'in development': 'construction_plus',
                # mothballed
                'mothballed': 'mothballed_plus',
                'idle': 'mothballed_plus',
                'shut in': 'mothballed_plus',
                # retired
                'retired': 'retired_plus',
                'closed': 'retired_plus',
                'decommissioned': 'retired_plus',})

    # make sure all statuses align with no space rule
    gdf_map_ready['status'] = gdf_map_ready['status'].apply(lambda x: x.strip().replace(' ','-'))
    gdf_map_ready['status_legend'] = gdf_map_ready['status_legend'].apply(lambda x: x.strip().replace('_','-'))

    print(set(gdf_map_ready['status'].to_list()))
    gdf_map_ready_no_status = gdf_map_ready[gdf_map_ready['status']== '']
    gdf_map_ready= gdf_map_ready[gdf_map_ready['status']!= '']
    # print(len(gdf_map_ready))
    gdf_map_ready_no_status.to_csv(f'{path_for_test_results}no-status-africa-energy.csv')
    gdf_map_ready = gdf_map_ready.fillna('')
    print(set(gdf_map_ready['status'].to_list()))
    gdf_map_ready['status'] = gdf_map_ready['status'].apply(lambda x: x.lower())
    print(set(gdf_map_ready['status'].to_list()))

    return gdf_map_ready

africa_df_status_update = map_ready_statuses(africa_df_converted)

def map_ready_countries(gdf):

    gdf_map_ready = gdf.copy()
    # just need to make sure all countries are separated by a comma and have a comma after last country as well
    # GOGET has a hyphen in countries
    # GOIT has comma separated in countries
    # hydropower has two columns country1 and country2
    # GGIT has comma separated in countries
    # grouped_tracker_before = gdf.groupby('tracker', as_index=False)['id'].count()
    # print(f'In map ready before adjustment: {grouped_tracker_before}')
    # # no return

    # gdf_map_ready['area2'] = gdf_map_ready['area2'].fillna('')
    for row in gdf_map_ready.index:

        if gdf_map_ready.loc[row, 'area2'] != '':
            gdf_map_ready.loc[row, 'areas'] = f"{gdf_map_ready.loc[row, 'area'].strip()};{gdf_map_ready.loc[row, 'area2'].strip()};"
            print(f"Found a area2! Hydro? {gdf_map_ready.loc[row, 'areas']} {gdf_map_ready.loc[row, 'tracker']} {gdf_map_ready.loc[row, 'name']}")

        else:
            # make it so all areas even just one end with a comma 
            gdf_map_ready.loc[row, 'areas'] = f"{gdf_map_ready.loc[row, 'areas'].strip()};"
    
    # grouped_tracker_after = gdf.groupby('tracker', as_index=False)['id'].count()

    # print(f'In map ready after adjustment: {grouped_tracker_after}')
    

    return gdf_map_ready

africa_df_country_update = map_ready_countries(africa_df_status_update)

# Define a function to check for valid values
def is_valid_goget(x):
    return isinstance(x, (int, float)) and not pd.isna(x) and x != '' and x != 0 and x != 0.0


def workarounds_eg_interim(gdf):
    # copy tracker column to be full display name not custom 
    gdf = gdf.copy()
    # create column that has capacity with appended units 
    
    # for items with no capacity remove capacity
    # goget, and GCMT need to not have anything in capacity
    
    # make capacity == ''
    # production_mask = gdf[gdf['tracker']=='GOGET' or 'GCMT'] 
    for row in gdf.index:
        prod_oil = gdf.loc[row, 'prod_oil']
        prod_gas = gdf.loc[row, 'prod_gas']
        cap = gdf.loc[row, 'capacity']
        tracker = (gdf.loc[row, 'tracker'])
        #if goget then make capacity table and capacity details empty
        
        if tracker == 'GOGET':
            gdf.loc[row, 'capacity-table'] = ''
            gdf.loc[row, 'capacity-details'] = ''
            
            if is_valid_goget(prod_oil):
                # round it then if either is '' we can remove it later we filter on it before adding to table or details
                prod_oil = str(float(round(prod_oil, 3)))

                gdf.loc[row, 'prod-oil-table'] = f'{prod_oil} million bbl/y'
                gdf.loc[row, 'prod-oil-details'] = f'{prod_oil} (million bbl/y)'
                
            if is_valid_goget(prod_gas):
                prod_gas = str(float(round(prod_gas, 3)))

                gdf.loc[row, 'prod-gas-table'] = f'{prod_gas} million m³/y'
                gdf.loc[row, 'prod-gas-details'] = f'{prod_gas} (million m³/y)'

                
    # in that column above make it so all units within a project are summed and it's called total capacityDONE
    
    # handle more than one country for sat image say intranational project
    # if there is more than one ; you know it's a multiple country situation
    # so we want to say in areas-display "multi country"
    # we would also want to overwrite the subnat and say nothing ""
    gdf['count_of_semi'] = gdf.apply(lambda row: row['areas'].split(';')) # if len of list is more than 2, then split more than once
    gdf['multi-country'] = gdf.apply(lambda row: 't' if len(row['count_of_semi']) > 2 else 'f')
    print(gdf['multi-country'])
    # if t then make areas-display 
    gdf['areas-subnat-sat-display'] = gdf.apply(lambda row: f"{row['subnat'],row['areas']}" if row['multi-country'] == 'f' else 'Multiple Countries/Areas')


    # here we want to get the total number of units at the project level and put it at each row
    # so however it gets pulled into details it is shown
    # units_per_project = gdf.groupby(['name', 'status'], as_index=False)['geometry'].count()
    # print(len(f'length of units_per_project series from groupby: {units_per_project}'))
    # print(units_per_project.columns)
<<<<<<< HEAD
=======

>>>>>>> 81808373619ec7caaa2f0f4b2814984378a66e84
    # gdf = pd.merge(left=gdf, right=units_per_project, on='name', how='outer')
    
    #### figure out 
    
    # tracker type not clickable! 
    
    #### nice to have
    
    # also create a country multiple or not column for displaying on map's sat picture  
    # escape apostrophe and / 
    
    return gdf
    
workarounds_eg_interim_gdf = workarounds_eg_interim(africa_df_country_update)

def last_min_fixes(gdf):
    gdf = gdf.copy()
    print(len(gdf))
    gdf.drop_duplicates(inplace=True)
    print(len(gdf))

    gdf['name'] = gdf['name'].fillna('')
    gdf['url'] = gdf['url'].fillna('')
    print(gdf.columns)
    # TODO for now remove drop to understand why all GOGET empty is now not being filled and also prod doesn't come through
    # gdf = gdf.drop(['original_units','conversion_factor', 'area2', 'region2', 'subnat2', 'capacity1', 'capacity2', 'Latitude', 'Longitude'], axis=1) # 'cleaned_cap'
    # handle countries/areas ? 
    # gdf['areas'] = gdf['areas'].apply(lambda x: x.replace('Guinea,Bissau','Guinea-Bissau')) # reverse this one special case back 
    print(f'Missing areas: {gdf[gdf["areas"]==""]}') # check if missing!!   
    # handle situation where Guinea-Bissau IS official and ok not to be split into separate countries 
    gdf['areas'] = gdf['areas'].apply(lambda x: x.replace('Guinea,Bissau','Guinea-Bissau')) 
    
    # something is happening when we concat, we lose goget's name ... 
    # gdf_empty_name = gdf[gdf['name']=='']
    print(f"This is cols for gdf: {gdf.columns}")
    gdf['wiki-from-name'] = gdf.apply(lambda row: f"https://www.gem.wiki/{row['name'].strip().replace(' ', '_')}", axis=1)
    # row['url'] if row['url'] != '' else 
    gdf['url'] = gdf.apply(lambda row: row['wiki-from-name'] if row['url'] == '' else row['url'], axis=1)
    
    # gdf['name'] = gdf.apply(lambda row: row['name'] if row['name'] != '' else row['url'].split('www.gem.wiki/')[1].replace('_', ' '))

    gdf_empty_url = gdf[gdf['url'] == '']
    
    # print(gdf_empty_url[['tracker', 'wiki-from-name', 'url']])
    
    # assign Africa to all regions
    gdf['region'] = 'Africa'
    
    gdf.columns = [col.replace('_', '-') for col in gdf.columns]    
    # let's also look into empty url, by tracker I can assign a filler
    
    # gdf['url'] = gdf['url'].apply(lambda row: row[filler_url] if row['url'] == '' else row['url'])
    
    # gdf['capacity'] = gdf['capacity'].apply(lambda x: x.replace('--', '')) # can't do that ... 
    
    # translate acronyms to full names 
    
    gdf['tracker-display'] = gdf['tracker-custom'].map(tracker_to_fullname)
    gdf['tracker-legend'] = gdf['tracker-custom'].map(tracker_to_legendname)
    # print(set(gdf['tracker-display'].to_list()))
    # print(f'Figure out capacity dtype: {gdf.dtypes}')
    gdf['capacity'] = gdf['capacity'].apply(lambda x: str(x).replace('--', ''))
    # gdf['capacity'] = gdf['capacity'].apply(lambda x: x.replace('', pd.NA))
    # gdf['capacity'] = gdf['capacity'].astype(float) # Stuck with concatting like strings for now? ValueError: could not convert string to float: ''
    
    # REMOVE TODO not found!! 
    
     
    return gdf


africa_gdf_country_update = last_min_fixes(workarounds_eg_interim_gdf)


def final_count(gdf):
    grouped_tracker = gdf.groupby('tracker', as_index=False)['id'].count()
    print(print(grouped_tracker))
    # no return

final_count(africa_gdf_country_update)

# RE RUN with fixed new GOGET OR 

# call all

# prep_df = create_prep_file(prep_file_key, prep_file_tab)

# prep_dict = df_to_dict(prep_df)

# loop_over_dict(prep_dict)

# list_of_dfs = create_all_dfs(prep_dict)

# list_of_renamed_dfs = rename_drop_all_dfs(list_of_dfs, prep_dict)

# one_df = combine_all_dfs(list_of_renamed_dfs)

# africa_df = filter_by_country_all_dfs(one_df)

# df_with_point_geom = convert_all_coords(africa_df)
# africa_gdf = create_geojson(df_with_point_geom)
# # africa_gdf = ''

# concatted_gdf = concat_exisiting_geojson(goit_geojson, ggit_geojson, ggit_pipes_geojson, africa_gdf)

# # africa_gdf = analyze_geojson('/Users/gem-tah/Desktop/GEM_INFO/GEM_WORK/earthrise-maps/gem_tracker_maps/trackers/africa-energy/compilation_output/2024-06-20/africa_energy_tracker_2024-06-20.geojson')

# africa_gdf_converted = capacity_conversions(africa_gdf) 

# africa_gdf_status_update = map_ready_statuses(africa_gdf_converted)
# print(set(africa_gdf_status_update['tracker'].to_list()))

# africa_gdf_country_update = map_ready_countries(africa_df_status_update)
# africa_gdf_country_update.to_csv(f'africa_gdf_country_update{today_date}.csv')

# gdf_to_geojson(africa_gdf_converted, f'{path_for_download_and_map_files}africa_gdf_converted.geojson')

gdf_to_geojson(africa_gdf_country_update, f'{path_for_download_and_map_files}{geojson_file_of_all_africa}')


# # df = pd.DataFrame(africa_gdf_country_update)
# # trackers = set(df['tracker'].to_list())
# # for row in df.index:
# #     if pd.isna(df.at[row, 'capacity'])==False:
# #     # remove empty capacity
# #         for tracker in trackers:
# #             print(df[df['tracker']== tracker]['scaling_capacity'])
# # df['capacity'] = df['capacity'].fillna('')
# # print(df.columns)
# # df_africa = df[df['capacity']!='']
# # for tracker in trackers:
# #     print(df_africa[df_africa['tracker']== tracker]['scaling_capacity'])
# #     print(df_africa[df_africa['tracker']== tracker]['capacity'])
# #     print(df_africa[df_africa['tracker']== tracker]['conversion_factor'])
africa_gdf_country_update.to_excel(f'{path_for_download_and_map_files}{geojson_file_of_all_africa.split(".geojson")[0]}.xlsx', index=False)


